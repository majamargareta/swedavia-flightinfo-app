from openpyxl import load_workbook
import json

# Excel file containing official destination statistics from Swedavia.
EXCEL_FILE = "destinations-statistik-2025.xlsx"

# Open the workbook and select the active worksheet.
workbook = load_workbook(EXCEL_FILE)
sheet = workbook.active

# Dictionary used to store city-country mappings.
city_country = {}

# Read destination data from the Excel file.
# The data starts on row 9 in the statistics sheet.
for row in sheet.iter_rows(min_row=9, values_only=True):
    country = row[0]
    city = row[1]

    # Skip empty rows.
    if not country or not city:
        continue

    # Skip summary rows in the statistics file.
    if "Summa" in str(country):
        continue

    # Add city-country mapping from the Excel data.
    city_country[city] = country.title()


# Manual mappings for destinations that are missing from
# the statistics file or use different names in the API.
manual_mappings = {
    "Arvidsjaur": "Sweden",
    "Gällivare": "Sweden",
    "Göteborg": "Sweden",
    "Hemavan Tärnaby": "Sweden",
    "Kalmar": "Sweden",
    "Karlstad": "Sweden",
    "Kiruna": "Sweden",
    "Luleå": "Sweden",
    "Malmö": "Sweden",
    "Ronneby": "Sweden",
    "Skellefteå": "Sweden",
    "Sundsvall": "Sweden",
    "Sveg": "Sweden",
    "Torsby": "Sweden",
    "Umeå": "Sweden",
    "Vilhelmina": "Sweden",
    "Visby": "Sweden",
    "Ängelholm": "Sweden",
    "Åre Östersund": "Sweden",

    "London LHR": "Great Britain",
    "London LGW": "Great Britain",
    "London STN": "Great Britain",
    "Paris CDG": "France",
    "Paris ORY": "France",
    "Rome FCO": "Italy",
    "Milan LIN": "Italy",
    "Milan MXP": "Italy",
    "New York JFK": "United States",
    "New York EWR": "United States",
    "Istanbul IST": "Turkey",
    "Istanbul SAW": "Turkey",
    "Warsaw WMI": "Poland",
    "Tokyo HND": "Japan",
    "Addis Abeba": "Ethiopia",
    "Alicante": "Spain",
    "Barcelona": "Spain",
    "Bergamo": "Italy",
    "Bergen": "Norway",
    "Bucharest OTP": "Romania",
    "Charleroi": "Belgium",
    "Cologne": "Germany",
    "Heraklion": "Greece",
    "Munich": "Germany",
    "Trapani": "Italy",
    "Tromso": "Norway",
    "Vaasa": "Finland",
    "Vilnius": "Lithuania"
}

# Merge manual mappings with the generated data.
city_country.update(manual_mappings)

# Save the final city-country mapping to a JSON file.
with open(
    "city_country.json",
    "w",
    encoding="utf-8"
) as file:
    json.dump(
        city_country,
        file,
        ensure_ascii=False,
        indent=4
    )

# Print a summary when the file has been created.
print(f"Created city_country.json with {len(city_country)} cities")